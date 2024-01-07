#-------------------------------------------------------------------------------
# Name:        module1
# Purpose:
#
# Author:      LENOVO
#
# Created:     28-12-2023
# Copyright:   (c) LENOVO 2023
# Licence:     <your licence>
#-------------------------------------------------------------------------------

#-------------------------------------------------------------------------------
# Name:        main
# Purpose:
#
# Author:      LENOVO
#
# Created:     03-11-2023
# Copyright:   (c) LENOVO 2023
# Licence:     <your licence>
#-------------------------------------------------------------------------------

from kivymd.app import MDApp
from kivy.lang import Builder
from kivy.core.window import Window
from kivymd.uix.dialog import MDDialog
from kivy.app import App
import openpyxl
from openpyxl import Workbook
import pathlib
from kivymd.uix.datatables import MDDataTable
import pandas as pd
from kivy.metrics import dp
from kivymd.uix.button import MDRaisedButton

#Window.size = (450, 720)

KV='''
Screen:
    MDBoxLayout:
        orientation: 'vertical'
        padding: 16
        spacing:5

        MDBoxLayout:
            padding: 20
            spacing: 10
            pos_hint: {'center_y': 0.5}
            size_hint_y: 0.3

            MDTextField:
                hint_text: "Name"
                mode: "rectangle"
                id: name
                icon_right: "account-outline"
                width: 200
                font_size: 48
                pos_hint: {"center_x": 0.5}
                size_hint_y: 1.2
                size_hint_x: 2

            MDTextField:
                hint_text: "Standard"
                mode: "rectangle"
                id: standard
                icon_right: "list-box-outline"
                width: 200
                font_size: 48
                size_hint_y: 1.2
                size_hint_x: 2

            MDRaisedButton:
                text: "Add"
                md_bg_color: "#333333"
                size_hint_y: 1.08
                on_press: app.add()

        MDCard:
            size_hint: None, None
            size: 1005, 1700
            pos_hint: {'center_x': 0.5, 'center_y': 0.5}
            padding: 15
            spacing: 25

            MDScrollView:
                id: card_layout
                do_scroll_x: False

        MDBoxLayout:
            padding: 20
            spacing: 10
            pos_hint: {'center_y': 0.5}
            size_hint_y: 0.3
            MDRaisedButton:
                text: "Search"
                icon_right: "database-search-outline"
                size_hint_y: 1.2
                size_hint_x: 2
                md_bg_color: "#333333"
                pos_hint: {'center_y': 0.5}
                on_press: app.search()
            MDRaisedButton:
                id: reset_button
                text: "Reset Table"
                icon_right: "refresh"
                size_hint_y: 1.2
                size_hint_x: 2
                md_bg_color: "#333333"
                pos_hint: {'center_y': 0.5}
                on_press: app.reset_table()
'''


class dataEntry(MDApp):
    def build(self):
        self.theme_cls.theme_style = "Dark"
        self.theme_cls.primary_palette = "DeepPurple"
        return Builder.load_string(KV)

    def add(self, *args):
        # Your add method code here
        name = self.root.ids.name.text
        standard = self.root.ids.standard.text

        if name == '' or standard == '':
            #messagebox.showwarning('warning', 'Fill all fields')
            print("fill all fields")
        else:
            print(name)
            print(standard)

            file = openpyxl.load_workbook('Backend_data.xlsx')
            sheet = file.active
            sheet.cell(column=1, row=sheet.max_row+1, value=name)
            sheet.cell(column=2, row=sheet.max_row, value=standard)

            file.save(r'Backend_data.xlsx')
            #messagebox.showinfo('info', 'Data Added')
            self.reset_table()

            self.root.ids.name.text = ''
            self.root.ids.standard.text=''

    def reset_search_buttons(self, search_button_visible=True):
        # Your reset_search_buttons method code here
        """
        Show or hide the search and reset buttons based on the search_button_visible parameter.
        """
        self.root.ids.search_button.opacity = 1 if search_button_visible else 0
        self.root.ids.reset_button.opacity = 1 if not search_button_visible else 0

    def search(self):
        # Your search method code here
        name = self.root.ids.name.text
        standard = self.root.ids.standard.text

        if name == '' and standard == '':
            # If both input boxes are empty, show a message or handle it as needed
            print("Enter at least one value for search")
        else:
            data = pd.read_excel('Backend_data.xlsx')

            if name:
                data = data[data['Name'].str.contains(name, case=False, na=False)]
            if standard:
                data = data[data['Standard'].astype(str).str.contains(standard, case=False, na=False)]

            cols = data.columns.values
            values = data.values

            table = MDDataTable(
                pos_hint={'center_x': 0.2, 'center_y': 0.2},
                column_data=[(col, dp(35)) for col in cols],
                row_data=values,
                use_pagination=True,
                padding=10,
                pagination_menu_pos='auto',
                rows_num=10,
            )

            #Add a refresh button to call the reset_table function
            #refresh_button = MDRaisedButton(
            #    text="Refresh",
            #    md_bg_color="red",
             #   size_hint_y=0.2,
             #   on_press=self.reset_table
            #)

            # Add the table and refresh button to the card layout
            self.root.ids.card_layout.clear_widgets()
            self.root.ids.card_layout.add_widget(table)
            #self.root.ids.card.add_widget(refresh_button)


    def on_start(self):
        # Your on_start method code here
        file = pathlib.Path('Backend_data.xlsx')
        if file.exists():
            pass
        else:
            file = Workbook()
            sheet = file.active
            sheet['A1'] = "Name"
            sheet['B1'] = "Standard"
            file.save('Backend_data.xlsx')

        self.reset_table()

        #load_data_table()

        #self.root.ids.datatable.add_widget(
        #    use_pagination=True,
        #    column_data=[
        #        (col, dp(30))
        #        for col in cols
        #    ],
        #    row_data=values
        #)



        # Clear the existing widgets in the MDScrollView

    def reset_table(self):
        # Your reset_table method code here
        data = pd.read_excel('Backend_data.xlsx')
        cols = data.columns.values
        values = data.values

        table = MDDataTable(
            pos_hint={'center_x': 0.2, 'center_y': 0.2},
            #size_hint=(1, 0.6),
            column_data=[(col, dp(35)) for col in cols],
            row_data=values,
            #size_hint=(1.7, 1),
            use_pagination=True,
            padding = 10,
            pagination_menu_pos = 'auto',
            rows_num = 10,

        )
        self.root.ids.card_layout.clear_widgets()
        self.root.ids.card_layout.add_widget(table)

if __name__ == "__main__":
    data_app = dataEntry()
    data_app.run()
